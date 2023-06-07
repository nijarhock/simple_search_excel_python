from openpyxl import load_workbook
import re
import os
import json
import ast

# set constant
current_working_directory = os.getcwd()
excel_file = ""
rules = {}

# set translate
dictionary = {
  "less_than" : "<",
  "equal" : "=",
  "not_equal" : "!=",
  "less_than_equal" : "<=",
  "greater_then" : ">",
  "greater_then_equal" : ">=",
  "like" : "Mengandung Kata",
  "not_like" : "Tidak Mengandung Kata"
}
while True:

  # main menu
  print("Selamat datang di program pencarian sederhana excel")
  print("Silahkan pilih menu berikut : ")
  print("\n")
  print("1. ", "Pilih Excel")
  print("2. ", "Pilih Rule")
  print("3. ", "Proses Search")
  print("4. ", "Exit")
  command = input("pilih menu: ")

  if command == "1":
    print("\n")
    # filter just excel file
    dir_list = os.listdir(current_working_directory)
    filtered_list = [file for file in dir_list if ".xlsx" in file]
    filtered_list = [file for file in filtered_list if ".xlsx#" not in file]

    # show excel file
    for index, f in enumerate(filtered_list):
      print(index+1, ". ", f)
    
    print(len(filtered_list)+1, ". ", "Kembali")
    choose_file = input("Pilih file: ")

    # if file choosing
    if int(choose_file) == len(filtered_list)+1:
      continue
    else:
      excel_file = filtered_list[int(choose_file)-1]
      continue
  elif command == "2":
    print("\n")
    print("Rule file must have extension rule")
    print("Example : ")
    print("{")
    print('\t"0" : { ')
    print('\t\t"column" : "Harga Campaign"')
    print('\t\t"operator" : "less_than"')
    print('\t\t"value" : 50000')
    print('\t} ')
    print("}")
    print("Note : ")
    for key in dictionary:
      print(" - ", key, " = ", dictionary[key])

    dir_list = os.listdir(current_working_directory)
    filtered_list = [file for file in dir_list if ".rule" in file]
    # show rule file
    for index, f in enumerate(filtered_list):
      print(index+1, ". ", f)

    print(len(filtered_list)+1, ". ", "Kembali")
    choose_file = input("Pilih file: ")

    # if file choosing
    if int(choose_file) == len(filtered_list)+1:
      continue
    else:
      f = open(filtered_list[int(choose_file)-1], "r")
      rules = ast.literal_eval(f.read())
      continue
    
  elif command == "3":
    # check already choosing file
    if len(rules) == 0:
      print("Belum pilih File rule")
      continue
    if excel_file == "":
      print("Belum pilih File excel")
      continue
    # load file excel
    try:
      wb = load_workbook(filename = excel_file)
      ws = wb['0']
    except:
      print("Error cek kembali file excel")
      exit()
    # get max length column and max row
    try:
      max_col = len(tuple(ws.columns))
      max_row = len(tuple(ws.rows))
    except:
      print("Error cek kembali file excel")
      exit()
    # set start row and checking row
    start_row = 2
    checking_row = 2
    list_row = []

    # function
    def search_column(rules, woorkbook, max_col):
      column = 0
      for col in woorkbook.iter_cols(min_row=1, max_col=max_col, max_row=1):
        for cell in col:
          if cell.value == rules['column']:
            rules['column_number'] = column
            break
          column += 1
      return rules

    def search_row(rule, i, value):
      if i == rule['column_number']:
        if rule['operator'] == 'less_than':
          if int(value) < rule['value']:
            return True
        elif rule['operator'] == 'equal':
          if int(value) == rule['value']:
            return True
        elif rule['operator'] == 'not_equal':
          if int(value) != rule['value']:
            return True
        elif rule['operator'] == 'less_than_equal':
          if int(value) <= rule['value']:
            return True
        elif rule['operator'] == 'greater_then':
          if int(value) > rule['value']:
            return True
        elif rule['operator'] == 'greater_then_equal':
          if int(value) >= rule['value']:
            return True
        elif rule['operator'] == 'like':
          if re.search(rule['value'], str(value)) != None:
            return True
        elif rule['operator'] == 'not_like':
          if re.search(rule['value'], str(value)) == None:
            return True
            
      return False
    # end function

    # searching coordinate column 
    try:
      for rule in rules:
        rules[rule] = search_column(rules[rule], ws, max_col)
    except:
      print("Error cek kembali file rule")
      exit()

    # searching row by rules
    try:
      for row in ws.iter_rows(min_row=start_row, max_col=max_col, max_row=max_row, values_only=True):
        i = 0
        count_true = 0
        for value in row:
          if value is None:
            break
          for rule in rules:
            if search_row(rules[rule], i, value):
              count_true += 1 
          i += 1
        if count_true == len(rules):
          list_row.append(checking_row) 
        checking_row += 1
    except:
      print("Error cek kembali file excel")
      exit()
    # show list row
    print("Row yang terfilter dengan rule sebagai berikut : ")
    for rule in rules:
      print("- ", rules[rule]["column"], dictionary[rules[rule]["operator"]], rules[rule]["value"])
    print("\n")
    print("No Row ")
    for l in list_row:
      print(l)
  else:
    exit()

